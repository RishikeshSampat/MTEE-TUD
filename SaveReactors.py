import os
import numpy
from pyevtk.hl import*
from openpyxl import Workbook

def SaveData(list,name):
    loc = os.getcwd()
    dest_filename = loc + '/%s%i.xlsx' %(name,1)
    wb = Workbook()

    # General data
    ws = wb.active
    ws.title = 'General'
    row=0
    for i in list:
        row += 1
        ws.cell(column=1, row=row, value=i)

    wb.save(filename=dest_filename)


def Save(Reactors, header, data, MassImb, OutletReact, InletReact):
    loc=os.getcwd()
    dest_filename = loc + '/data%i.xlsx' % 1
    wb = Workbook()

    # General data
    ws = wb.active
    ws.title = 'General'
    NO_ind=0
    NO2_ind=0
    O2_ind=0
    H2O_ind=0

    col = 1
    row = 1
    ws.cell(column=col, row=row, value='Reactor name')
    row += 1
    ws.cell(column=col, row=row, value='Temperature [K]')
    row += 1
    ws.cell(column=col, row=row, value='Pressure [Pa]')
    row += 1
    ws.cell(column=col, row=row, value='Residence time [ms]')
    row += 1
    ws.cell(column=col, row=row, value='Mass [kg]')
    row += 1
    ws.cell(column=col, row=row, value='Volume [m^3]')
    row += 1
    ws.cell(column=col, row=row, value='Density [kg/m^3]')
    row += 1
    ws.cell(column=col, row=row, value='Mean Molecular Weight []')
    for sp in Reactors[0].thermo.species_names:
        row += 1
        ws.cell(column=col, row=row, value=sp)
        if sp=='NO':
            NO_ind=row
        if sp=='NO2':
            NO2_ind=row
        if sp == 'O2':
            O2_ind = row
        if sp=='H2O':
            H2O_ind=row


    Pressure=[]
    Volume=[]
    Temperature=[]
    Mass=[]
    O2=[]
    NO=[]
    NO2=[]
    H2O=[]

    col = 2
    for r in Reactors:
        row = 1
        ws.cell(column=col, row=row, value=r.name)
        row += 1
        ws.cell(column=col, row=row, value=r.T)
        Temperature.append(r.T)
        row += 1
        ws.cell(column=col, row=row, value=r.thermo.P)
        Pressure.append(r.thermo.P)
        row += 1
        ws.cell(column=col, row=row, value=0)
        row += 1
        ws.cell(column=col, row=row, value=r.mass)
        Mass.append(r.mass)
        row += 1
        ws.cell(column=col, row=row, value=r.volume)
        Volume.append(r.volume)
        row += 1
        ws.cell(column=col, row=row, value=r.density)
        row += 1
        ws.cell(column=col, row=row, value=r.thermo.mean_molecular_weight)
        for x in r.thermo.X:
            row += 1
            ws.cell(column=col, row=row, value=x)
            if row == NO_ind:
                 NO.append(x)
            if row == NO2_ind:
                 NO2.append(x)
            if row == O2_ind:
                 O2.append(x)
            if row == H2O_ind:
                 H2O.append(x)
        col += 1

    wb.save(filename=dest_filename)
    x = header.index('X-Coordinate') + 1
    y = header.index('Y-Coordinate') + 1
    z = header.index('Z-Coordinate') + 1

    X = numpy.array(data[19][x], dtype=numpy.float64)
    Y = numpy.array(data[19][y], dtype=numpy.float64)
    Z = numpy.array(data[19][z], dtype=numpy.float64)
    dat = {}
    for sp in Reactors[0].thermo.species_names:
        temp=[]
        index=Reactors[0].thermo.species_names.index(sp)
        for r in Reactors:
            temp.append(r.thermo.X[index])
        dat[sp]=numpy.array(temp, dtype=numpy.float64)
    temp1=[]
    temp2=[]
    temp3=[0]*len(Reactors)
    temp4=[0]*len(Reactors)
    for r in Reactors:
        temp1.append(r.T)
        temp2.append(r.thermo.P)
    for i in InletReact:
        temp3[i]=1
    for o in OutletReact:
        temp4[o]=1
    dat['Temperature'] = numpy.array(temp1, dtype=numpy.float64)
    dat['Pressure'] = numpy.array(temp2, dtype=numpy.float64)
    dat['MassImbalance'] = numpy.array(MassImb, dtype=numpy.float64)
    dat['Inlet'] = numpy.array(temp3, dtype=numpy.float64)
    dat['Outlet'] = numpy.array(temp4, dtype=numpy.float64)
    pointsToVTK("para_CRN", X, Y, Z, data=dat)
    NOx=[]
    R=8314
    perc_corr=0.15#O2 correction
    for conc in range(len(NO)):
        P=Pressure[conc]
        T=Temperature[conc]
        V=Volume[conc]
        N_wet=P*V/(R*T)
        N_dry=N_wet*(1-H2O[conc])
        NO_dry=NO[conc]*N_wet/N_dry
        #O2 correction
        N_O2=O2[conc]*N_wet
        O2_corr=(perc_corr/(1-perc_corr))*(N_dry-N_O2)
        N_corr=N_dry-N_O2+O2_corr
        NO_corr=NO_dry*N_dry/N_corr
        NOx.append(NO_corr*1e-6)
    z2plot=[]
    NOxplot=[]
    MassReact=[]
    for ind in range(len(data[19][z])):
        zcoord=data[19][z][ind]
        if zcoord in z2plot:
            z_id=z2plot.index(zcoord)
            NOxplot[z_id]=(NOxplot[z_id]*MassReact[z_id]+NOx[ind]*Mass[ind])/(MassReact[z_id]+Mass[ind])
            MassReact[z_id]=MassReact[z_id]+Mass[ind]
        else:
            z2plot.append(zcoord)
            NOxplot.append(NOx[ind])
            MassReact.append(Mass[ind])

    dest_filename = loc + '/Emissions.xlsx'
    wb = Workbook()

    # General data
    ws = wb.active
    ws.title = 'NOx'
    row=1
    ws.cell(column=1, row=row, value='Z(mm)')
    ws.cell(column=2, row=row, value='NOx ppm')
    row+=1
    for j in range(len(z2plot)):
        ws.cell(column=1, row=row, value=z2plot[j])
        ws.cell(column=2, row=row, value=NOxplot[j])
        row+=1
    import matplotlib.pyplot as plt
    plt.figure(3)
    plt.plot(z2plot,NOxplot)
    plt.title('Emissions')
    plt.xlabel('Z (mm)')
    plt.ylabel('NOx (dry volume ppm @ 15% O2)')
    plt.show()

    #write NOx data to file
    Emissions={'NOx':NOx,'Z':z2plot}