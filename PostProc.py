from openpyxl import Workbook, load_workbook
import math
import pickle
import numpy
from pyevtk.hl import*
print "Reading Data files"
with open("graph2plot.pkl", 'rb') as f:
    graph = pickle.load(f)
with open("header.pkl", 'rb') as f:
    header = pickle.load(f)
with open("data.pkl", 'rb') as f:
    data = pickle.load(f)
theta_index=header.index('Abs. Angular Coordinate')+1
theta_max=max(data[19][theta_index])
theta_min=min(data[19][theta_index])
print theta_index
print theta_max
print theta_min
#define normal of plane ax+by+cz=0
a=-1
b=2.414#1.4966
c=0
# format of cell2plot:
# [psi, z, temp, pressure,...]
cells2plot={}
x_ind=header.index('X-Coordinate')+1
y_ind=header.index('Y-Coordinate')+1
z_ind=header.index('Z-Coordinate')+1
Queue=[]
diction={'X':[],'Y':[],'Z':[],'Reactor':[]}
print len(data[19][z_ind])
for i in range(len(data[19][z_ind])):
    x=data[19][x_ind][i]
    y=data[19][y_ind][i]
    z=data[19][z_ind][i]
    resid=a*x+b*y+c*z
    if abs(resid)<1e-5:
        cells2plot[i+1]=[math.sqrt(x**2+y**2), z]
        Queue.append(i+1)
        

wb=load_workbook('data1.xlsx')
ws=wb['General']
print len(cells2plot)
print "Searching cell"
for number in range(ws.max_row-1):
    val=ws.cell(row=number+2,column=1)
    title=str(val.value)
    title=title.replace(" ","")
    
    diction[title]=[]

for reactor in graph:
    cells = graph[reactor]['cells']
    for i in cells:
        if i in Queue:
            Queue.remove(i)
            for row in range(ws.max_row-1):
                r=row+2
                val=ws.cell(row=r,column=reactor+1)
                cells2plot[i].append(val.value)
                val_title=ws.cell(row=r,column=1)
                title=str(val_title.value)
                title=title.replace(" ","")
                diction[title].append(val.value)
            diction['X'].append(cells2plot[i][0])
            diction['Y'].append(cells2plot[i][1])
            diction['Z'].append(0)
            diction['Reactor'].append(reactor)
    if Queue==[]:
        break
"""while Queue!=[]:
    elem = Queue[0]
    for reactor in graph:
        cells=graph[reactor]['cells']
        if elem in cells:
            Queue.remove(Queue[0])
            for row in range(1):
                val=ws.cell(row=row+2,column=reactor+1)
                cells2plot[elem].append(val.value)
            break"""

print "Writing file"
"""with open('postproc.dat','w') as p:
    p.write('Title="Post Processing"\n')
    p.write('VARIABLES= "X", "Y"')
    for row in range(ws.max_row - 1):
        val = ws.cell(row=row + 2, column=1)
        head=str(val.value)
        p.write(', "%s"' %head.replace(" ",""))
    p.write('\n')
    indices=cells2plot.keys()
    numpts=len(cells2plot)
    p.write('Zone T="Frame 0", I=%i, J=%i\n'%(numpts,numpts))
    for element in cells2plot:
        for x in cells2plot[element]:
            if cells2plot[element].index(x)==0:
                p.write(str(x))
                
            else:
                p.write(' '+str(x))
                
                    
                 
        p.write('\n')"""

print len(diction['X'])
print len(diction['Y'])
print len(diction['Z'])
print len(diction['Temperature[K]'])
for key in diction.keys():
    diction[key]=numpy.array(diction[key], dtype=numpy.float64)

pointsToVTK("PostProc", diction['X'], diction['Y'], diction['Z'], data=diction)

